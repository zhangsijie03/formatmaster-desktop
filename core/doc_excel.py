"""Excel/CSV/XLS 格式转换"""

import io
import os
import tempfile


class DocExcelMixin:
    """Excel/CSV/XLS 转换方法"""

    # ========== Excel (.xlsx) 转换 ==========

    def _xlsx_to_pdf(self, inp, out, cb):
        """XLSX → PDF：优先办公软件保真导出，缺失时使用中文内置引擎。

        旧实现直接走 ReportLab 默认 Helvetica，中文会变成方框或 ``IIII``；
        统一进入 office PDF 引擎后可注册 CJK 字体，并保留可搜索文本层。
        """
        if cb: cb(30, "准备导出PDF...")
        from core.doc_office_pdf import excel_to_pdf
        try:
            ok, engine = excel_to_pdf(inp, out, cb)
            if cb:
                cb(100, f"转换完成（{engine}）")
            return ok
        except Exception as exc:
            if cb:
                cb(-1, f"错误: {str(exc)[:200]}")
            return False

    def _xlsx_to_csv(self, inp, out, cb):
        import openpyxl
        import csv
        wb = openpyxl.load_workbook(inp, data_only=True)
        ws = wb.active
        with open(out, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            total = ws.max_row or 1
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if self._cancel: return False
                writer.writerow(row)
                if cb:
                    cb(20 + int(i * 70 / total), f"写入第{i+1}行...")
        return True

    def _xlsx_to_txt(self, inp, out, cb):
        import openpyxl
        wb = openpyxl.load_workbook(inp, data_only=True)
        ws = wb.active
        lines = []
        for row in ws.iter_rows(values_only=True):
            lines.append('\t'.join(str(c) if c is not None else '' for c in row))
        self._write_text(out, '\n'.join(lines))
        return True

    def _xlsx_to_image(self, inp, out, cb):
        if cb: cb(20, "读取表格...")
        import openpyxl
        from PIL import Image, ImageDraw, ImageFont
        wb = openpyxl.load_workbook(inp, data_only=True)
        ws = wb.active
        data = self._make_table_data(ws)
        try:
            font = ImageFont.truetype("msyh.ttc", 16)
        except Exception:
            font = ImageFont.load_default()
        cell_w, cell_h = 120, 30
        col_w = [cell_w] * max(len(data[0]) if data else 1, 1)
        for row_idx, row in enumerate(data):
            for col_idx, val in enumerate(row):
                est = len(str(val)) * 10 + 20
                if col_idx < len(col_w) and est > col_w[col_idx]:
                    col_w[col_idx] = min(est, 300)
        img_w = max(400, sum(col_w) + 20)
        img_h = max(200, len(data) * cell_h + 40)
        img = Image.new('RGB', (img_w, img_h), 'white')
        draw = ImageDraw.Draw(img)
        draw.rectangle([0, 0, img_w-1, img_h-1], outline='#cccccc')
        y = 10
        for row_idx, row in enumerate(data):
            x = 10
            if row_idx == 0:
                draw.rectangle([0, y, img_w, y+cell_h], fill='#4472C4')
            for col_idx, val in enumerate(row):
                cw = col_w[col_idx] if col_idx < len(col_w) else cell_w
                fill = '#D6E4F0' if row_idx == 0 else ('white' if row_idx % 2 == 1 else '#F2F2F2')
                draw.rectangle([x, y, x+cw, y+cell_h], fill=fill, outline='#cccccc')
                text_color = 'white' if row_idx == 0 else 'black'
                draw.text((x+4, y+6), str(val)[:30], fill=text_color, font=font)
                x += cw
            y += cell_h
        img.save(out)
        if cb: cb(100, "转换完成")
        return True

    def _xlsx_to_html(self, inp, out, cb):
        if cb: cb(20, "读取表格...")
        import openpyxl
        wb = openpyxl.load_workbook(inp, data_only=True)
        ws = wb.active
        rows_html = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            tag = 'th' if i == 0 else 'td'
            cells = ''.join(f'<{tag}>{self._safe_html(str(c) if c is not None else "")}</{tag}>' for c in row)
            rows_html.append(f'<tr>{cells}</tr>')
        body = f'<table border="1" cellpadding="6" style="border-collapse:collapse;font-size:12px">{"".join(rows_html)}</table>'
        self._write_text(out, self._build_html_page(body))
        if cb: cb(100, "转换完成")
        return True

    def _xlsx_to_md(self, inp, out, cb):
        if cb: cb(20, "读取表格...")
        import openpyxl
        wb = openpyxl.load_workbook(inp, data_only=True)
        ws = wb.active
        lines = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = [str(c) if c is not None else '' for c in row]
            lines.append('| ' + ' | '.join(vals) + ' |')
            if i == 0:
                lines.append('| ' + ' | '.join(['---'] * len(vals)) + ' |')
        self._write_text(out, '\n'.join(lines))
        if cb: cb(100, "转换完成")
        return True

    # ========== Excel97 (.xls) 转换 ==========

    def _xls_to_xlsx(self, inp, out, cb):
        if cb: cb(20, "读取Excel97...")
        import xlrd
        import openpyxl
        wb_old = xlrd.open_workbook(inp)
        ws_old = wb_old.sheet_by_index(0)
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        for r in range(ws_old.nrows):
            if self._cancel: return False
            for c in range(ws_old.ncols):
                cell = ws_old.cell(r, c)
                ws_new.cell(row=r+1, column=c+1, value=cell.value)
            if cb and r % 50 == 0:
                cb(20 + int(r * 70 / max(ws_old.nrows, 1)), f"第{r+1}/{ws_old.nrows}行...")
        wb_new.save(out)
        return True

    def _xls_to_pdf(self, inp, out, cb):
        """XLS 先转为 XLSX，再复用统一 CJK PDF 引擎。"""
        fd, xlsx_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            if not self._xls_to_xlsx(inp, xlsx_path, cb):
                return False
            return self._xlsx_to_pdf(xlsx_path, out, cb)
        finally:
            try:
                os.remove(xlsx_path)
            except OSError:
                pass

    def _xls_to_csv(self, inp, out, cb):
        import xlrd
        import csv
        wb = xlrd.open_workbook(inp)
        ws = wb.sheet_by_index(0)
        with open(out, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            for r in range(ws.nrows):
                if self._cancel: return False
                row = [str(ws.cell(r, c).value) if ws.cell(r, c).value is not None else '' for c in range(ws.ncols)]
                writer.writerow(row)
                if cb and r % 50 == 0:
                    cb(20 + int(r * 70 / max(ws.nrows, 1)), f"第{r+1}/{ws.nrows}行...")
        return True

    def _xls_to_txt(self, inp, out, cb):
        import xlrd
        wb = xlrd.open_workbook(inp)
        ws = wb.sheet_by_index(0)
        lines = []
        for r in range(ws.nrows):
            row = '\t'.join(str(ws.cell(r, c).value) if ws.cell(r, c).value is not None else '' for c in range(ws.ncols))
            lines.append(row)
        self._write_text(out, '\n'.join(lines))
        return True

    def _xls_to_image(self, inp, out, cb):
        if cb: cb(20, "读取Excel97...")
        import xlrd
        from PIL import Image, ImageDraw, ImageFont
        wb = xlrd.open_workbook(inp)
        ws = wb.sheet_by_index(0)
        data = []
        for r in range(ws.nrows):
            row = [str(ws.cell(r, c).value) if ws.cell(r, c).value is not None else '' for c in range(ws.ncols)]
            data.append(row)
        if not data:
            data = [['(空表格)']]
        try:
            font = ImageFont.truetype("msyh.ttc", 16)
        except Exception:
            font = ImageFont.load_default()
        cell_w, cell_h = 120, 30
        col_w = [cell_w] * max(len(data[0]), 1)
        for row_idx, row in enumerate(data):
            for col_idx, val in enumerate(row):
                est = len(str(val)) * 10 + 20
                if col_idx < len(col_w) and est > col_w[col_idx]:
                    col_w[col_idx] = min(est, 300)
        img_w = max(400, sum(col_w) + 20)
        img_h = max(200, len(data) * cell_h + 40)
        img = Image.new('RGB', (img_w, img_h), 'white')
        draw = ImageDraw.Draw(img)
        y = 10
        for row_idx, row in enumerate(data):
            x = 10
            if row_idx == 0:
                draw.rectangle([0, y, img_w, y+cell_h], fill='#4472C4')
            for col_idx, val in enumerate(row):
                cw = col_w[col_idx] if col_idx < len(col_w) else cell_w
                fill = '#D6E4F0' if row_idx == 0 else ('white' if row_idx % 2 == 1 else '#F2F2F2')
                draw.rectangle([x, y, x+cw, y+cell_h], fill=fill, outline='#cccccc')
                text_color = 'white' if row_idx == 0 else 'black'
                draw.text((x+4, y+6), str(val)[:30], fill=text_color, font=font)
                x += cw
            y += cell_h
        img.save(out)
        if cb: cb(100, "转换完成")
        return True

    def _xls_to_html(self, inp, out, cb):
        if cb: cb(20, "读取Excel97...")
        import xlrd
        wb = xlrd.open_workbook(inp)
        ws = wb.sheet_by_index(0)
        rows_html = []
        for r in range(ws.nrows):
            cells = ''.join(f'<{"th" if r==0 else "td"}>{self._safe_html(str(ws.cell(r,c).value) if ws.cell(r,c).value is not None else "")}</{"th" if r==0 else "td"}>' for c in range(ws.ncols))
            rows_html.append(f'<tr>{cells}</tr>')
        body = f'<table border="1" cellpadding="6" style="border-collapse:collapse;font-size:12px">{"".join(rows_html)}</table>'
        self._write_text(out, self._build_html_page(body))
        if cb: cb(100, "转换完成")
        return True

    def _xls_to_md(self, inp, out, cb):
        if cb: cb(20, "读取Excel97...")
        import xlrd
        wb = xlrd.open_workbook(inp)
        ws = wb.sheet_by_index(0)
        lines = []
        for r in range(ws.nrows):
            vals = [str(ws.cell(r, c).value) if ws.cell(r, c).value is not None else '' for c in range(ws.ncols)]
            lines.append('| ' + ' | '.join(vals) + ' |')
            if r == 0:
                lines.append('| ' + ' | '.join(['---'] * len(vals)) + ' |')
        self._write_text(out, '\n'.join(lines))
        return True

    # ========== CSV 转换 ==========

    def _csv_rows(self, inp):
        """GBK/UTF-8 等 CSV 统一解码，同时保留引号内换行。"""
        import csv
        return csv.reader(io.StringIO(self._read_text(inp), newline=""))

    def _csv_to_pdf(self, inp, out, cb):
        """CSV 通过 XLSX 中间层复用中文与分页能力更完整的 PDF 引擎。"""
        fd, xlsx_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            if not self._csv_to_xlsx(inp, xlsx_path, cb):
                return False
            return self._xlsx_to_pdf(xlsx_path, out, cb)
        finally:
            try:
                os.remove(xlsx_path)
            except OSError:
                pass

    def _csv_to_html(self, inp, out, cb):
        if cb: cb(20, "读取CSV...")
        rows_html = []
        for i, row in enumerate(self._csv_rows(inp)):
            cells = ''.join(f'<{"th" if i==0 else "td"}>{self._safe_html(c)}</{"th" if i==0 else "td"}>' for c in row)
            rows_html.append(f'<tr>{cells}</tr>')
        body = f'<table border="1" cellpadding="6" style="border-collapse:collapse">{"".join(rows_html)}</table>'
        self._write_text(out, self._build_html_page(body))
        return True

    def _csv_to_xlsx(self, inp, out, cb):
        if cb: cb(20, "读取CSV...")
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for i, row in enumerate(self._csv_rows(inp)):
            if self._cancel: return False
            ws.append(row)
            if cb and i % 100 == 0:
                cb(20 + min(70, i // 10), f"写入第{i+1}行...")
        wb.save(out)
        if cb: cb(100, "转换完成")
        return True

    def _csv_to_md(self, inp, out, cb):
        if cb: cb(20, "读取CSV...")
        lines = []
        for i, row in enumerate(self._csv_rows(inp)):
            vals = [c for c in row]
            lines.append('| ' + ' | '.join(vals) + ' |')
            if i == 0:
                lines.append('| ' + ' | '.join(['---'] * len(vals)) + ' |')
        self._write_text(out, '\n'.join(lines))
        return True
